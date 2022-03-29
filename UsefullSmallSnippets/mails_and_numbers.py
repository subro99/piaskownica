import sys
import pyperclip
import re
from traceback import print_tb
from pathlib import Path
from openpyxl import load_workbook, Workbook

'''Utworzenie pliku excel lub otworzenie jeśli już istnieje o takiej nazwie'''
try:
    wb = load_workbook(filename="mails_numbers.xlsx")
    print(f"arkusze:{wb.sheetnames}")
    if "maile" not in wb.sheetnames:   
        ws=wb.create_sheet("maile")
    if "numery" not in wb.sheetnames:
        ws1=wb.create_sheet("numery")
    else:
        print()
        ws = wb["maile"]
        ws1 = wb["numery"]
        print(type(wb.sheetnames))
except:
    wb=Workbook()
    ws=wb.active
    ws.title= "maile" #tytul 1 arkusza
    ws1=wb.create_sheet("numery") # utworzenie i tytuł 2 arkusza


text = pyperclip.paste()  # przypisanie skopiowanego tekstu do zmiennej

'''Wyszukanie patternu maila i dodanie do excela'''
mailRegex = re.compile("\S+@\S+")
mails = mailRegex.findall(text)
print(f"Znalezniono {len(mails)} adresy e-mail:")
for mail in mails:
    print(mail)
    ws.append([mail])
print()

'''Wyszukanie patternu numeru tel i dodanie do excela'''
phoneRegex = re.compile("\+?\d+[\d\s\-\.]{8,13}")
phones = phoneRegex.findall(text)
print(f"Znaleziono {len(phones)} numery telefonu:")
for phone in phones:
    print(phone)
    ws1.append([phone])

wb.save(filename="mails_numbers.xlsx")
